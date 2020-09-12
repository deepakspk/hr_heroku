from django.shortcuts import render, redirect
from django.utils import timezone
from django.urls import reverse_lazy
from django.views.generic import (View, ListView, TemplateView, DetailView, CreateView, UpdateView, DeleteView)
from . forms import EmployeeForm, DepartmentForm, PayslipForm, ProcessSalaryForm, AdditionalForm, DeductionForm, TimesheetForm, LabourContractForm, WorkPermitForm, EidCollectForm, NewEidForm, PassportReturnUpdateForm
from . import models
import re, os
from datetime import date
from django.db.models import Sum, Avg, Max, Count, Min
from django.db.models.functions import Coalesce
from django.db.models import Value
from openpyxl import Workbook
from openpyxl.styles import colors, Color, Border, Side, PatternFill, Font, GradientFill, Alignment
from os import path
from django.http import HttpResponse
from easy_pdf.views import PDFTemplateView, PDFTemplateResponseMixin, TemplateResponseMixin
from easy_pdf.rendering import render_to_pdf_response
from django.contrib import messages
import csv, io
from .utils import getPdf
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.http import FileResponse
from django.http import Http404
from django.views.generic.detail import BaseDetailView
from django.db.models import Max
from django.db.models import Q


# Create your views here.
class EmployeeListView(ListView):
    template_name = 'payroll/employee_list.html'
    context_object_name = 'employee_list'
    
    def get_queryset(self):
        queryset= models.Employee.objects.all()
        if "employeeid" in self.request.GET:
            if self.request.GET["employeeid"] !="":            
                queryset= models.Employee.objects.filter(employee_id=self.request.GET["employeeid"])
        if "hotel" in self.request.GET:
            if self.request.GET["hotel"] !="":            
                queryset= queryset.filter(hotel=self.request.GET["hotel"])
        if "status" in self.request.GET:
            if self.request.GET["status"] !="":            
                queryset= queryset.filter(employee_status=self.request.GET["status"])              
        return queryset


    def get_context_data(self, *args, **kwargs):
        context = super(EmployeeListView,self).get_context_data(*args,**kwargs)
        
        hotel = models.Hotel.objects.all()       
        context['hotel'] = hotel

        status = models.Employee.objects.all().values_list('employee_status',flat=True).distinct()
        
        context['status'] = status


        return context


class EmployeeDetailView(DetailView):
    model = models.Employee
    context_object_name = 'employee_detail'
    template_name='payroll/employee_detail.html'

    def get_context_data(self, *args, **kwargs):
        data = super(EmployeeDetailView,self).get_context_data(*args,**kwargs)
        data['labourcontract'] = models.LaborContract.objects.filter(employee=self.object)
        data['workpermit'] = models.WorkPermit.objects.filter(employee=self.object)
        data['eidcollect'] = models.EidCollect.objects.filter(employee=self.object)
        data['neweid'] = models.NewEid.objects.filter(employee=self.object)
        return data

class EmployeeCreateView(CreateView):
    template_name = 'payroll/employee_create.html'
    form_class = EmployeeForm

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

class EmployeeUpdateView(UpdateView):
    model = models.Employee
    form_class = EmployeeForm
    template_name='payroll/employee_update.html'


class EmployeeDeleteView(DeleteView):
    model = models.Employee
    success_url = reverse_lazy("payroll:employee_list")

    ### Start of Additional View ###
class AdditionalListView(ListView):
    template_name = 'payroll/additional_list.html'
    context_object_name = 'additional_list'
    model = models.Additionalpay

    def get_context_data(self, *args, **kwargs):
        context = super(AdditionalListView,self).get_context_data(*args,**kwargs)
        list = context['additional_list']
        totalParticular = list.all().count()
        totalAdd = list.aggregate(Sum('amount'))['amount__sum']
        tot_add = (round(totalAdd,2))
        context['tot_add'] = tot_add
        context['totalParticular'] = totalParticular
        return context

class AdditionalDetailView(DetailView):
    model = models.Additionalpay
    context_object_name = 'additional_detail'
    template_name='payroll/additional_detail.html'

    def get_context_data(self, *args, **kwargs):
        data = super(AdditionalDetailView,self).get_context_data(*args,**kwargs)
        return data

class AdditionalCreateView(CreateView):
    template_name = 'payroll/additional_create.html'
    form_class = AdditionalForm
    success_url = reverse_lazy("payroll:additional_list")

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

class AdditionalUpdateView(UpdateView):
    model = models.Additionalpay
    form_class = AdditionalForm
    template_name='payroll/additional_update.html'
    success_url = reverse_lazy("payroll:additional_list")

class AdditionalDeleteView(DeleteView):
    model = models.Additionalpay
    success_url = reverse_lazy("payroll:additional_list")
    ### End of  Additional View ###

### Start of Deductions View ###
class DeductionListView(ListView):
    template_name = 'payroll/deduction_list.html'
    context_object_name = 'deduction_list'
    model = models.Deduction

    def get_context_data(self, *args, **kwargs):
        context = super(DeductionListView,self).get_context_data(*args,**kwargs)
        list = context['deduction_list']
        totalParticular = list.all().count()
        totalded = list.aggregate(Sum('amount'))['amount__sum']
        tot_ded = (round(totalded,2))
        context['tot_ded'] = tot_ded
        context['totalParticular'] = totalParticular
        return context

class DeductionDetailView(DetailView):
    model = models.Deduction
    context_object_name = 'deduction_detail'
    template_name='payroll/deduction_detail.html'

def get_context_data(self, *args, **kwargs):
    data = super(DeductionDetailView,self).get_context_data(*args,**kwargs)
    return data

class DeductionCreateView(CreateView):
    template_name = 'payroll/deduction_create.html'
    form_class = DeductionForm
    success_url = reverse_lazy("payroll:deduction_list")

def get(self, request, *args, **kwargs):
    form = self.form_class()
    return render(request, self.template_name, {'form': form})

class DeductionUpdateView(UpdateView):
    model = models.Deduction
    form_class = DeductionForm
    template_name='payroll/deduction_update.html'
    success_url = reverse_lazy("payroll:deduction_list")

class DeductionDeleteView(DeleteView):
    model = models.Deduction
    success_url = reverse_lazy("payroll:deduction_list")

### End of  Deductions View ###

### Start of Timesheet View ###
class EmployeeTimesheet(ListView):
    template_name = 'payroll/timesheet_list.html'
    context_object_name = 'timesheet_list'
    model = models.Timesheet

    def get_context_data(self, *args, **kwargs):
        context = super(EmployeeTimesheet,self).get_context_data(*args,**kwargs)
        return context

class EmployeeTimesheetCreateView(CreateView):
    template_name = 'payroll/timesheet_create.html'
    form_class = TimesheetForm
    success_url = reverse_lazy("payroll:timesheet_list")

def get(self, request, *args, **kwargs):
    form = self.form_class()
    return render(request, self.template_name, {'form': form})

class EmployeeTimesheetUpdateView(UpdateView):
    model = models.Timesheet
    form_class = TimesheetForm
    template_name='payroll/timesheet_update.html'
    success_url = reverse_lazy("payroll:timesheet_list")

class EmployeeTimesheetDeleteView(DeleteView):
    model = models.Timesheet
    success_url = reverse_lazy("payroll:timesheet_list")

def timesheet_upload(request):
    # declaring template
    template = "payroll/timesheet_upload.html"
    data = models.Timesheet.objects.all()
    # prompt is a context variable that can have different values      depending on their context
    prompt = {
        'order': 'Order of the CSV should be Employee, Month, partial, Normal, Weekend, Holidays, Vacation, Sick, LWOP, Total',
        'timesheets': data
              }
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        if column[0] != "":
            getemp = models.Employee.objects.get(employee_id=column[0])
            created = models.Timesheet.objects.create(
            employee=getemp,
            month=column[1],
            partial=column[2],
            normal=column[3],
            weekend=column[4],
            holiday=column[5],
            vacation=column[6],
            sick=column[7],
            lwop=column[8],
            total=column[9],
            )
    context = {}
    return redirect("payroll:timesheet_list")

### End of Timesheet View ###

def employee_upload(request):
    # declaring template
    template = "payroll/employee_upload.html"
    data = models.Employee.objects.all()
    # prompt is a context variable that can have different values      depending on their context
    prompt = {
        'order': 'Order of the CSV should be Employee id, first_name, middle_name, last_name, joining_date, gender, DOB eg.2020-5-19, religion, citizenship, birth country, email, phone, uae address, home country phone, home country address, designation, department',
        'employees': data
              }
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        if column[0] != "":
            getdep = models.Department.objects.get(department=column[9])
            gethotel = models.Hotel.objects.get(hotel=column[10])
            created = models.Employee.objects.create(
            employee_id         = column[0],
            first_name          = column[1],
            joining_date        = column[2],
            gender              = column[3],
            DOB                 = column[4],
            citizenship         = column[5],
            email               = column[6],
            phone               = column[7],
            designation         = column[8],
            department          = getdep,
            hotel               = gethotel,
            employee_status     = column[11]           
            )
    context = {}
    return redirect("payroll:employee_list")

### End of Bulk employee upload ###

def paydetail_upload(request):
    # declaring template
    template = "payroll/paydetail_upload.html"
    data = models.Clinical.objects.all()
    # prompt is a context variable that can have different values      depending on their context
    prompt = {
        'order': 'Order of the CSV should be Employee id , basic_salary, housing_allowance, transportation_allowance, other, work_permit, cost_center, mode_of_payment, bank_ac, payment_method, employee_unique, agent_id, work_permit_8_digits, personal_no_14_digits, bank_name_routing' ,
        'paydetail': data
              }
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        if column[0] != "":
            getemp = models.Employee.objects.get(employee_id=column[0])
            getct = models.CostCenter.objects.get(cost_center=column[6])
            created = models.Payslip.objects.create(
            employee               =   getemp,
            basic_salary           =   column[1],
            housing_allowance      =   column[2],
            transportation_allowance = column[3],
            other                  =   column[4],
            work_permit            =   column[5],
            cost_center            =   getct,
            mode_of_payment        =   column[7],
            bank_ac                =   column[8],
            payment_method         =   column[9],
            employee_unique        =   column[10],
            agent_id               =   column[11],
            work_permit_8_digits   =   column[12],
            personal_no_14_digits  =   column[13],
            bank_name_routing      =   column[14],
            )
    context = {}
    return redirect("payroll:payslip_list")
# ### End of Bulk paydetail upload ###

def clinical_upload(request):
    # declaring template
    template = "payroll/clinical_upload.html"
    data = models.Clinical.objects.all()
    # prompt is a context variable that can have different values      depending on their context
    prompt = {
        'order': 'Order of the CSV should be Employee id , basic_salary, housing_allowance, transportation_allowance, other, work_permit, cost_center, mode_of_payment, bank_ac, payment_method, employee_unique, agent_id, work_permit_8_digits, personal_no_14_digits, bank_name_routing' ,
        'paydetail': data
              }
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        if column[0] != "":

            getemp = models.Employee.objects.get(employee_id=column[0])
            created = models.Clinical.objects.create(
            employee=getemp,
            seha_id                =   column[1],
            zoho_id                =   column[2],
            company_email          =   column[3],
            source                 =   column[4],
            source_entity          =   column[5],
            category               =   column[6],
            hiring_status          =    column[7],
            emirates_id             =   column[8],
            Passport_no             =   column[9],
            passport_expiry         =   column[10],
            licence_authority       =   column[11],
            doh_licence_type        =   column[12],
            doh_licence_no          =   column[13],
            previous_company        =   column[14],
            licence_issued_date     =   column[15],
            licence_expiry_date     =   column[16],
            non_doh_licence_no      =   column[17],
            higest_qualification    =   column[18],
            date_qualified          =   column[19],
            clinic_exp_in_5_year    =   column[20],
            healthy_fit             =   column[21],
            competencies           =    column[22],
            visa_status            =    column[23],
            previous_visa_sponser  =    column[24],
            doh_licence_as         =    column[25],
            work_order             =    column[26],    
            )
    context = {}
    return redirect("payroll:employee_list")

# ### End of Bulk paydetail upload ###


def additional_upload(request):
    # declaring template
    template = "payroll/additional_upload.html"
    data = models.Additionalpay.objects.all()
    # prompt is a context variable that can have different values      depending on their context
    prompt = {
        'order': 'Order of the CSV should be Employee id, additional, additional label, amount, occur, start date eg: 2020-05-19, end date date eg: 2020-05-30',
        'paydetail': data
              }
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        if column[0] != "":
            getemp = models.Employee.objects.get(employee_id=column[0])
            pay = models.Payslip.objects.get(employee=getemp)
            created = models.Additionalpay.objects.create(
            payslip                =   pay,
            additional             =   column[1],
            additional_label       =   column[2],
            amount                 =   column[3],
            occur                  =   column[4],
            start_date             =   column[5],
            end_date               =   column[6],
            )
    context = {}
    return redirect("payroll:additional_list")
#### End of Bulk additional pay upload ###

def deduction_upload(request):
    # declaring template
    template = "payroll/deduction_upload.html"
    data = models.Deduction.objects.all()
    # prompt is a context variable that can have different values      depending on their context
    prompt = {
        'order': 'Order of the CSV should be Employee id, additional, additional label, amount, occur, start date eg: 2020-05-19, end date date eg: 2020-05-30',
        'paydetail': data
              }
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        if column[0] != "":
            getemp = models.Employee.objects.get(employee_id=column[0])
            pay = models.Payslip.objects.get(employee=getemp)
            created = models.Deduction.objects.create(
            payslip                =   pay,
            deduction             =   column[1],
            deduction_label       =   column[2],
            amount                 =   column[3],
            occur                  =   column[4],
            start_date             =   column[5],
            end_date               =   column[6],
            )
    context = {}
    return redirect("payroll:deduction_list")
#### End of Bulk additional pay upload ###

class DepartmentListView(ListView):
    template_name = 'payroll/department_list.html'
    context_object_name = 'department_list'
    model = models.Department

    def get_context_data(self, *args, **kwargs):
        context = super(DepartmentListView,self).get_context_data(*args,**kwargs)
        dep_list = context['department_list']
        empData = []
        for dep in dep_list:
            total_em = models.Employee.objects.filter(department=dep).count()
            department_head = None
            try:
                department_head = models.Employee.objects.get(department_head=True, department=dep)
            except models.Employee.DoesNotExist:
                department_head = "-"
            total_emp = total_em or 0
            empData.append([dep,total_emp,department_head])
        context['Data'] = empData

        return (context)

class DepartmentDetailView(DetailView):
    model = models.Department
    context_object_name = 'department_detail'

    template_name='payroll/department_detail.html'

    def get_context_data(self, *args, **kwargs):
        data = super(DepartmentDetailView,self).get_context_data(*args,**kwargs)
        return data

class DepartmentCreateView(CreateView):
    template_name = 'payroll/department_create.html'
    form_class = DepartmentForm
    success_url = reverse_lazy("payroll:department_list")

class DepartmentUpdateView(UpdateView):
    model = models.Department
    form_class = DepartmentForm
    template_name='payroll/department_update.html'
    success_url = reverse_lazy("payroll:department_list")



class DepartmentDeleteView(DeleteView):
    model = models.Department
    success_url = reverse_lazy("payroll:department_list")

class PayslipListView(ListView):
    template_name = 'payroll/payslip_list.html'
    context_object_name = 'payslip_list'

    def get_queryset(self):
        return models.Payslip.objects.all()

    def get_context_data(self, *args, **kwargs):
        context = super(PayslipListView,self).get_context_data(*args,**kwargs)
        context['payslips'] = models.Payslip.objects.all()

        return context

class PayslipDetailView(DetailView):
    model = models.Payslip
    context_object_name = 'payslip_detail'
    template_name='payroll/payslip_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['addPay'] = models.Additionalpay.objects.filter(payslip=self.object)
        context['dedPay'] = models.Deduction.objects.filter(payslip=self.object)
        return context

class PayslipCreateView(CreateView):
    template_name = 'payroll/payslip_create.html'
    form_class = PayslipForm
    success_url = reverse_lazy("payroll:payslip_list")

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            newId = form.save(commit=True)
            addTypes = []
            dedTypes =[]
            for key in request.POST.keys():
                if key.startswith('apay_type'):
                    addTypes.append(key)
                if key.startswith('dpay_type'):
                    dedTypes.append(key)
            for addType in addTypes:
                getInt = re.findall(r'\d+',addType)[0]
                checkOccur = request.POST.get('aoccurs'+str(getInt))
                if checkOccur == "Monthly" and request.POST.get('apay_type'+str(getInt)) !="":
                    addPay = models.Additionalpay.objects.create(
                        payslip=newId,
                        additional=request.POST.get('apay_type'+str(getInt)),
                        additional_label=request.POST.get('alabel'+str(getInt)),
                        amount=request.POST.get('aamount'+str(getInt)),
                        occur=request.POST.get('aoccurs'+str(getInt)),
                        start_date=request.POST.get('astart_date'+str(getInt)),
                        end_date=request.POST.get('aend_date'+str(getInt)),
                    )
                elif checkOccur == "Once" and request.POST.get('apay_type'+str(getInt)) !="":
                    addPay = models.Additionalpay.objects.create(
                        payslip=newId,
                        additional=request.POST.get('apay_type'+str(getInt)),
                        additional_label=request.POST.get('alabel'+str(getInt)),
                        amount=request.POST.get('aamount'+str(getInt)),
                        occur=request.POST.get('aoccurs'+str(getInt))
                    )
            for dedType in dedTypes:
                getInt = re.findall(r'\d+',dedType)[0]
                checkOccur = request.POST.get('dpay_occur'+str(getInt))
                if checkOccur == "Monthly" and request.POST.get('dpay_type'+str(getInt)) !="":
                    addPay = models.Deduction.objects.create(
                        payslip=newId,
                        deduction=request.POST.get('dpay_type'+str(getInt)),
                        deduction_label=request.POST.get('dpay_label'+str(getInt)),
                        amount=request.POST.get('dpay_amount'+str(getInt)),
                        occur=request.POST.get('dpay_occur'+str(getInt)),
                        start_date=request.POST.get('dpay_start'+str(getInt)),
                        end_date=request.POST.get('dpay_end'+str(getInt)),
                    )
                elif checkOccur == "Once" and request.POST.get('dpay_type'+str(getInt)) !="":
                    addPay = models.Deduction.objects.create(
                        payslip=newId,
                        deduction=request.POST.get('dpay_type'+str(getInt)),
                        deduction_label=request.POST.get('dpay_label'+str(getInt)),
                        amount=request.POST.get('dpay_amount'+str(getInt)),
                        occur=request.POST.get('dpay_occur'+str(getInt)),
                    )
            return redirect("payroll:payslip_list")


class PayslipUpdateView(UpdateView):
    model = models.Payslip
    form_class = PayslipForm
    template_name='payroll/payslip_update.html'
    success_url = reverse_lazy("payroll:payslip_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['addPay'] = models.Additionalpay.objects.filter(payslip=self.object)
        context['dedPay'] = models.Deduction.objects.filter(payslip=self.object)
        return context

class PayslipDeleteView(DeleteView):
    model = models.Payslip
    success_url = reverse_lazy("payroll:payslip_list")

class ThanksPage(TemplateView):
    template_name = 'thanks.html'

class PayrollRegisterView(ListView):
    template_name = 'payroll/payroll_register_list.html'
    context_object_name = 'payroll_register_list'
    model = models.ProcessSalary

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args,**kwargs)
        processing = models.ProcessSalary.objects.filter(status="Processing")
        paid = models.ProcessSalary.objects.filter(status="Paid")
        processing_data = []
        paid_data = []
        total_nos = 0
        total_gro = 0
        actual_total_gross = 0
        gross_pay = 0
        tot_gro = 0
        tot_add_pay = 0
        tot_ded = 0
        gra_tot = 0

        for pr in processing:
            total_nos = pr.employee.all().count()
            processing_data.append([pr,total_nos])

        for prs in paid:
            total_nos = prs.employee_count
            paid_data.append([prs,total_nos])

            total_gross = paid.aggregate(Sum('gross_pay'))['gross_pay__sum']
            actual_total_gross = paid.aggregate(Sum('actual_gross_pay'))['actual_gross_pay__sum']
            total_additional_pay = paid.aggregate(Sum('Additional_pay'))['Additional_pay__sum']
            total_deduction = paid.aggregate(Sum('deduction'))['deduction__sum']
            grand_total = paid.aggregate(Sum('total'))['total__sum']
            total_gro = total_gross or 0
            actual_total_gross = actual_total_gross or 0
            total_add = total_additional_pay or 0
            total_ded = total_deduction or 0
            grand_tot = grand_total or 0
            tot_gro = (round(total_gro,2))
            actual_total_gross = (round(actual_total_gross,2))
            tot_add_pay = (round(total_add,2)) or 0
            tot_ded = (round(total_ded,2)) or 0
            gra_tot = (round(grand_tot,2)) or 0
        calc_data= [tot_gro,tot_add_pay,tot_ded,gra_tot,actual_total_gross]

        context['processing_data'] = processing_data
        context['paid_data'] = paid_data
        context['calc_data'] = calc_data
        return context


class ProcessSalaryListView(ListView):
    template_name = 'payroll/process_salary_list.html'
    context_object_name = 'process_salary_list'
    queryset = models.ProcessSalary.objects.filter(status="Processing")

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args,**kwargs)
        ps_list = context['process_salary_list']
        data = []
        total_nos = 0

        for ps in ps_list:
            total_nos = ps.employee.all().count()
            data.append([ps,total_nos])
        context['data'] = data
        return context

class ProcessSalaryDetailView(DetailView):
    model = models.ProcessSalary
    context_object_name = 'process_salary_detail'
    template_name='payroll/process_salary_detail.html'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status == "Paid":
            return redirect("payroll:process_salary_list")
        else:
            context = self.get_context_data(object=self.object)
            return self.render_to_response(context)

    def get_context_data(self, *args, **kwargs):
        data = super(ProcessSalaryDetailView,self).get_context_data(*args,**kwargs)
        getEmp = data['process_salary_detail'].employee.all()
        empData = []
        total_gross_pay = 0
        total_actualgross = 0
        total_additional_pay = 0
        total_deduction = 0
        total_lwop = 0
        total_net_pay = 0
        getgross = 0

        for emp in getEmp:
            empAdd = models.Additionalpay.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
            empDed = models.Deduction.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']

            try:
                getgross = models.Timesheet.objects.get(employee=emp.employee)
                # perhour = (emp.gross_pay/192)*hours
                # actualgross = hours*perhour           )
                actualgross = (emp.gross_pay/getgross.total)*getgross.pay_days
                lwop = (emp.gross_pay/getgross.total)*getgross.lwop
            except models.Timesheet.DoesNotExist:
                getgross = 0
                actualgross = 0
                lwop = getgross


            actualgross = (round(actualgross,2))
            empAddAmount = empAdd or 0
            empDedAmount = empDed or 0
            addamt = (round(empAddAmount,2))
            dedamt = (round(empDedAmount,2))
            lwop = (round(lwop,2))
            net = (emp.gross_pay + addamt) - dedamt - lwop
            total_gross_pay += emp.gross_pay
            total_actualgross += actualgross
            total_additional_pay += addamt
            total_deduction += dedamt
            total_lwop += lwop
            total_net_pay += net
            empData.append([emp,emp.gross_pay,addamt,dedamt, net, actualgross, getgross, lwop])
        empData2 = [total_gross_pay,total_additional_pay,total_deduction, total_net_pay,total_actualgross,total_lwop]
        data['empData'] = empData
        data['empData2'] = empData2
        return data

def update_ps(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    total_gross_pay = 0
    total_actual_gross_pay = 0
    total_additional_pay = 0
    total_deduction = 0
    total_net_pay = 0
    for emp in data.employee.all():
        empCount = data.employee.all().count()
        getemployeetimesheet = models.Timesheet.objects.get(employee=emp.employee)

        empAdd = models.Additionalpay.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deduction.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empAddAmount = empAdd or 0
        empDedAmount = empDed or 0
        actual_gross_pay = (emp.gross_pay/getemployeetimesheet.total)*getemployeetimesheet.pay_days
        actual_gross_pay = actual_gross_pay or 0
        net = (actual_gross_pay + empAddAmount) - empDedAmount
        total_gross_pay += emp.gross_pay
        total_actual_gross_pay += actual_gross_pay
        total_additional_pay += empAddAmount
        total_deduction += empDedAmount
        total_net_pay += net

        #create report for each employee
        getEmployee = models.Employee.objects.get(pk=emp.employee.pk)
        newReport = models.Report.objects.create(
                employee_id = getEmployee.employee_id,
                first_name = getEmployee.first_name,
                last_name = getEmployee.last_name,
                department = getEmployee.department.department,
                designation = getEmployee.designation,
                payroll_id = data.payroll_id,
                mode_of_payment = emp.mode_of_payment,
                work_permit = emp.work_permit,
                cost_center = emp.cost_center,
                payroll_month = data.salary_month,
                start_date    = data.start_date,
                finish_date   = data.finish_date,
                basic_salary = emp.basic_salary,
                housing_allowance = emp.housing_allowance,
                transportation_allowance = emp.transportation_allowance,
                other = emp.other,
                lowp_deduction = emp.gross_pay - actual_gross_pay,

            )

        newTimesheet = models.ReportTimesheet.objects.create(
            report = newReport,
            employee = getemployeetimesheet.employee,
            month = getemployeetimesheet.month,
            partial = getemployeetimesheet.partial,
            normal = getemployeetimesheet.normal,
            weekend = getemployeetimesheet.weekend,
            holiday = getemployeetimesheet.holiday,
            vacation = getemployeetimesheet.vacation,
            sick = getemployeetimesheet.sick,
            lwop = getemployeetimesheet.lwop,
            total = getemployeetimesheet.total,
        )
        # add additionl/deduction to the report
        getEmpAddPay = models.Additionalpay.objects.filter(payslip=emp)
        for empAddPay in getEmpAddPay:
            createAddPay = models.Additionalpays.objects.create(
                report = newReport,
                additional = empAddPay.additional,
                additional_label = empAddPay.additional_label,
                amount = empAddPay.amount,
            )
        getEmpDedPay = models.Deduction.objects.filter(payslip=emp)
        for empDedPay in getEmpDedPay:
            createAddPay = models.Deductions.objects.create(
                report = newReport,
                deduction = empDedPay.deduction,
                deduction_label = empDedPay.deduction_label,
                amount = empDedPay.amount,
            )
    data.employee_count = empCount
    data.gross_pay = total_gross_pay
    data.actual_gross_pay = total_actual_gross_pay
    data.Additional_pay = total_additional_pay
    data.deduction = total_deduction
    data.total = total_net_pay
    data.status = 'Paid'
    data.save()

    return redirect('payroll:process_salary_list')

class ProcessSalaryCreateView(CreateView):
    template_name = 'payroll/process_salary_create.html'
    form_class = ProcessSalaryForm

class ProcessSalaryUpdateView(UpdateView):
    model = models.ProcessSalary
    form_class = ProcessSalaryForm
    template_name='payroll/process_salary_update.html'

class ProcessSalaryDeleteView(DeleteView):
    model = models.ProcessSalary
    success_url = reverse_lazy("payroll:process_salary_list")

def EmployeePayslip(request, pk):
    data = models.Payslip.objects.get(pk=pk)
    emp_data = models.Employee.objects.get(pk=data.employee.pk)
    addpay = models. Additionalpay.objects.filter(payslip=data)
    dedpay = models. Deduction.objects.filter(payslip=data)

    empAdd = models.Additionalpay.objects.filter(payslip=data).aggregate(Sum('amount'))['amount__sum']
    empDed = models.Deduction.objects.filter(payslip=data).aggregate(Sum('amount'))['amount__sum']
    netPay = ( data.gross_pay + empAdd ) - empDed

    return render(request,'payroll/employee_payslip.html',
        {'data':data,
        'emp_data':emp_data,
        'empAdd':empAdd,
        'empDed':empDed,
        'netPay':netPay,
        'addpay':addpay,
        'dedpay':dedpay,
        })

def ListofEmpInDep(request, pk):
    data = models.Department.objects.get(pk=pk)
    list_employee = models.Employee.objects.filter(department=data)

    return render(request,'payroll/empdeplist.html',
        {'list_employee':list_employee,
        'data': data,
        })

class SalaryReport(ListView):
    template_name = 'payroll/salary_report.html'
    context_object_name = 'salary_report_list'
    model = models.Report

    def get_context_data(self, *args, **kwargs):
        context = super(SalaryReport,self).get_context_data(*args,**kwargs)
        list = context['salary_report_list']
        data = []

        for i in list:
            empAdd = models.Additionalpays.objects.filter(report=i).aggregate(Sum('amount'))['amount__sum']
            empDed = models.Deductions.objects.filter(report=i).aggregate(Sum('amount'))['amount__sum']
            empAddAmount = empAdd or 0
            empDedAmount = empDed or 0

            netPay = (i.gross_pay + empAddAmount) - empDedAmount
            data.append([i,empAddAmount, empDedAmount,netPay])
        context['data'] = data

        return context


class ReportDetailView(DetailView):
    model = models.Report
    context_object_name = 'report_detail'
    template_name='payroll/report_detail.html'

    def get_context_data(self, *args, **kwargs):
        data = super(ReportDetailView,self).get_context_data(*args,**kwargs)
        addpay = models.Additionalpays.objects.filter(report=data['object'])
        dedpay = models.Deductions.objects.filter(report=data['object'])
        empAdd = models.Additionalpays.objects.filter(report=data['object']).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deductions.objects.filter(report=data['object']).aggregate(Sum('amount'))['amount__sum']
        lowp = (data['object'].lowp_deduction) or 0
        empAddAmount = empAdd or 0
        empDedAmount = empDed or 0
        addamt = (round(empAddAmount,2))
        dedamt = (round(empDedAmount,2))
        dedamount = dedamt
        gross = data['object'].basic_salary + data['object'].housing_allowance + data['object'].transportation_allowance + data['object'].other - data['object'].lowp_deduction
        net = (gross + addamt) - dedamount
        yrsbasic = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('basic_salary'))['basic_salary__sum']

        yrshousing = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('housing_allowance'))['housing_allowance__sum']

        yrstransport = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('transportation_allowance'))['transportation_allowance__sum']

        yrsother = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('other'))['other__sum']

        yrslowp_deduction = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('lowp_deduction'))['lowp_deduction__sum']


        yrsGross = yrsbasic + yrshousing + yrstransport + yrsother - yrslowp_deduction
        yrsGross = (round(yrsGross,2))


        yrsAdd = models.Additionalpays.objects.filter(
                report__employee_id=data['object'].employee_id,
                report__payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('amount'))['amount__sum']
        yearAdd = yrsAdd or 0
        yearAdd = (round(yearAdd,2))


        yrsDed = models.Deductions.objects.filter(
                report__employee_id=data['object'].employee_id,
                report__payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('amount'))['amount__sum']

        yearDed = yrsDed or 0
        yrded = (round(yearDed,2))

        yrsNet = (yrsGross + yearAdd ) - yrded
        yearNet = yrsNet or 0
        yearNet = (round(yearNet,2))


        empData = [addamt,dedamount,net,addpay,dedpay,yearAdd,yrded,yrsGross,yearNet, gross]
        data['empData'] = empData
        return data


class ReportDetailView_Revised(DetailView):
    model = models.Report
    context_object_name = 'report_detail'
    template_name='payroll/report_detail_view.html'

    def get_context_data(self, *args, **kwargs):
        data = super(ReportDetailView_Revised,self).get_context_data(*args,**kwargs)

        addpay = models.Additionalpays.objects.filter(report=data['object'])
        dedpay = models.Deductions.objects.filter(report=data['object'])
        empAdd = models.Additionalpays.objects.filter(report=data['object']).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deductions.objects.filter(report=data['object']).aggregate(Sum('amount'))['amount__sum']
        timesheet = models.ReportTimesheet.objects.filter(report=data['object'])
        lowp = (data['object'].lowp_deduction) or 0

        basic_rate = (data['object'].basic_salary)/timesheet[0].total
        allowance_rate = (data['object'].housing_allowance)/timesheet[0].total
        transport_rate = (data['object'].transportation_allowance)/timesheet[0].total
        other_rate = (data['object'].other)/timesheet[0].total
        total_rate = basic_rate + allowance_rate  + transport_rate + other_rate

        normal_basic = basic_rate * timesheet[0].normal
        normal_allowance = allowance_rate * timesheet[0].normal
        normal_transport = transport_rate * timesheet[0].normal
        normal_other = other_rate * timesheet[0].normal
        normal_total = normal_basic + normal_allowance + normal_transport + normal_other

        weekend_basic = basic_rate * timesheet[0].weekend
        weekend_allowance = allowance_rate * timesheet[0].weekend
        weekend_transport = transport_rate * timesheet[0].weekend
        weekend_other = other_rate * timesheet[0].weekend
        weekend_total = weekend_basic + weekend_allowance + weekend_transport + weekend_other

        holiday_basic = basic_rate * timesheet[0].holiday
        holiday_allowance = allowance_rate * timesheet[0].holiday
        holiday_transport = transport_rate * timesheet[0].holiday
        holiday_other = other_rate * timesheet[0].holiday
        holiday_total = holiday_basic + holiday_allowance + holiday_transport + holiday_other

        vacation_basic = basic_rate * timesheet[0].vacation
        vacation_allowance = allowance_rate * timesheet[0].vacation
        vacation_transport = transport_rate * timesheet[0].vacation
        vacation_other = other_rate * timesheet[0].vacation
        vacation_total = vacation_basic + vacation_allowance + vacation_transport + vacation_other

        sick_basic = basic_rate * timesheet[0].sick
        sick_allowance = allowance_rate * timesheet[0].sick
        sick_transport = transport_rate * timesheet[0].sick
        sick_other = other_rate * timesheet[0].sick
        sick_total = sick_basic + sick_allowance + sick_transport + sick_other

        total_basic = normal_basic + weekend_basic + holiday_basic + vacation_basic + sick_basic
        total_allowance = normal_allowance + weekend_allowance + holiday_allowance + vacation_allowance + sick_allowance
        total_transport = normal_transport + weekend_transport + holiday_transport + vacation_transport + sick_transport
        total_other = normal_other + weekend_other + holiday_other + vacation_other + sick_other
        grand_total = total_basic + total_allowance + total_transport + total_other


        basic_rate = (round(basic_rate,2))
        allowance_rate = (round(allowance_rate,2))
        transport_rate = (round(transport_rate,2))
        other_rate = (round(other_rate,2))
        total_rate = (round(total_rate,2))

        normal_basic = (round(normal_basic,2))
        normal_allowance = (round(normal_allowance,2))
        normal_transport = (round(normal_transport,2))
        normal_other = (round(normal_other,2))
        normal_total = (round(normal_total,2))

        weekend_basic = (round(weekend_basic,2))
        weekend_allowance = (round(weekend_allowance,2))
        weekend_transport = (round(weekend_transport,2))
        weekend_other = (round(weekend_other,2))
        weekend_total = (round(weekend_total,2))

        holiday_basic = (round(holiday_basic,2))
        holiday_allowance = (round(holiday_allowance,2))
        holiday_transport = (round(holiday_transport,2))
        holiday_other = (round(holiday_other,2))
        holiday_total = (round(holiday_total,2))

        vacation_basic = (round(vacation_basic,2))
        vacation_allowance = (round(vacation_allowance,2))
        vacation_transport = (round(vacation_transport,2))
        vacation_other = (round(vacation_other,2))
        vacation_total = (round(vacation_total,2))

        sick_basic = (round(sick_basic,2))
        sick_allowance = (round(sick_allowance,2))
        sick_transport = (round(sick_transport,2))
        sick_other = (round(sick_other,2))
        sick_total = (round(sick_total,2))

        total_basic = (round(total_basic,2))
        total_allowance = (round(total_allowance,2))
        total_transport = (round(total_transport,2))
        total_other = (round(total_other,2))
        grand_total = (round(grand_total,2))

        empAddAmount = empAdd or 0
        empDedAmount = empDed or 0
        addamt = (round(empAddAmount,2))
        dedamt = (round(empDedAmount,2))
        dedamount = dedamt
        net = (grand_total + addamt) - dedamount
        yrsbasic = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('basic_salary'))['basic_salary__sum']

        yrshousing = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('housing_allowance'))['housing_allowance__sum']

        yrstransport = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('transportation_allowance'))['transportation_allowance__sum']

        yrsother = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('other'))['other__sum']

        yrslwop = models.Report.objects.filter(
                employee_id=data['object'].employee_id,
                payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('lowp_deduction'))['lowp_deduction__sum']

        yrsGross = yrsbasic + yrshousing + yrstransport + yrsother - yrslwop
        yrsGross = (round(yrsGross,2))

        yrsAdd = models.Additionalpays.objects.filter(
                report__employee_id=data['object'].employee_id,
                report__payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('amount'))['amount__sum']
        yearAdd = yrsAdd or 0
        yearAdd = (round(yearAdd,2))

        yrsDed = models.Deductions.objects.filter(
                report__employee_id=data['object'].employee_id,
                report__payroll_month__lte=data['object'].payroll_month
            ).aggregate(Sum('amount'))['amount__sum']

        yearDed = yrsDed or 0
        yrded = (round(yearDed,2))

        yrsNet = (yrsGross + yearAdd ) - yrded
        yearNet = yrsNet or 0
        yearNet = (round(yearNet,2))


        empData = [addamt,dedamount,net,addpay,dedpay,yearAdd,yrded,yrsGross,yearNet,basic_rate,allowance_rate,transport_rate,other_rate,total_rate,timesheet,
                    normal_basic,normal_allowance,normal_transport,normal_other,normal_total,
                    weekend_basic,weekend_allowance,weekend_transport,weekend_other,weekend_total,
                    holiday_basic,holiday_allowance,holiday_transport,holiday_other,holiday_total,
                    vacation_basic,vacation_allowance,vacation_transport,vacation_other,vacation_total,
                    sick_basic,sick_allowance,sick_transport,sick_other,sick_total,
                    total_basic, total_allowance, total_transport, total_other, grand_total, ]

        data['empData'] = empData
        return data



def ListofPayroll(request):
    data2 = models.ProcessSalary.objects.filter(status="Processing")
    data = models.ProcessSalary.objects.filter(status="Paid")
    datas2 = []
    datas = []


    for ps in data:
        emp_count = ps.employee.all().count()
        datas.append([ps,emp_count])
    payroll_count = data.count()
    total_gross = data.aggregate(Sum('gross_pay'))['gross_pay__sum']
    total_additional_pay = data.aggregate(Sum('Additional_pay'))['Additional_pay__sum']
    total_deduction = data.aggregate(Sum('deduction'))['deduction__sum']
    grand_total = data.aggregate(Sum('total'))['total__sum']
    total_gro = total_gross or 0
    total_add = total_additional_pay or 0
    total_ded = total_deduction or 0
    grand_tot = grand_total or 0
    tot_gro = (round(total_gro,2))
    tot_add_pay = (round(total_add,2)) or 0
    tot_ded = (round(total_ded,2)) or 0
    gra_tot = (round(grand_tot,2)) or 0

    for pay in data2:
        emp_count = pay.employee.all().count()
        datas2.append([pay,emp_count])
    payroll_count2 = data2.count()
    total_gross2 = data2.aggregate(Sum('gross_pay'))['gross_pay__sum']
    total_additional_pay2 = data2.aggregate(Sum('Additional_pay'))['Additional_pay__sum']
    total_deduction2 = data2.aggregate(Sum('deduction'))['deduction__sum']
    grand_total2 = data2.aggregate(Sum('total'))['total__sum']
    total_gro2 = total_gross2 or 0
    total_add2 = total_additional_pay2 or 0
    total_ded2 = total_deduction2 or 0
    grand_tot2 = grand_total2 or 0
    tot_gro2 = (round(total_gro2,2))
    tot_add_pay2 = (round(total_add2,2)) or 0
    tot_ded2 = (round(total_ded2,2)) or 0
    gra_tot2 = (round(grand_tot2,2)) or 0
    return render(request,'payroll/payroll_list.html',
        {'data':datas,
        'payroll_count':payroll_count,
        'tot_gro':tot_gro,
        'tot_add_pay':tot_add_pay,
        'tot_ded':tot_ded,
        'gra_tot':gra_tot,
        'data2':datas2,
        'payroll_count2':payroll_count2,
        'tot_gro2':tot_gro2,
        'tot_add_pay2':tot_add_pay2,
        'tot_ded2':tot_ded2,
        'gra_tot2':gra_tot2,
        })

def DetailofPayroll(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = models.Report.objects.filter(payroll_id=data.payroll_id)

    empCount = getEmp.count()
    empData = []
    total_gross = 0
    total_lwop = 0
    total_actual_gross = 0
    total_add = 0
    total_ded = 0
    total_net = 0

    for emp in getEmp:
        empAdd = models.Additionalpays.objects.filter(report=emp).aggregate(Sum('amount'))['amount__sum']
        empded = models.Deductions.objects.filter(report=emp).aggregate(Sum('amount'))['amount__sum']
        empAddAmount = empAdd or 0
        empdedAmount = empded or 0
        addamt=(round(empAddAmount,2))
        dedamt=(round(empdedAmount,2))
        actual_gross = emp.gross_pay - emp.lowp_deduction
        net = emp.gross_pay + addamt - dedamt - emp.lowp_deduction
        total_gross += emp.gross_pay
        total_lwop += emp.lowp_deduction
        total_actual_gross += actual_gross
        total_add += addamt
        total_ded += dedamt
        total_net += net
        empData.append([emp,emp.employee_id,emp.gross_pay,addamt,dedamt,net,actual_gross,emp.lowp_deduction])

    return render(request,'payroll/payroll_detail.html',
    {'data':data,
    'getEmp':getEmp,
    'empData':empData,
    'addamt':addamt,
    'empded':dedamt,
    'net':net,
    'empCount': empCount,
    'total_gross': total_gross,
    'total_actual_gross': total_actual_gross,
    'total_add': total_add,
    'total_ded': total_ded,
    'total_net': total_net,
    'total_lwop': total_lwop,
    })


def emppayslip(request,payrollid,empid):
    payslip = models.Report.objects.filter(employee_id=empid,payroll_id=payrollid )
    return redirect('/payroll/report_detail/'+str(payslip[0].pk))

def emppayslip_view(request,payrollid,empid):
    payslip = models.Report.objects.filter(employee_id=empid,payroll_id=payrollid )
    return redirect('/report_detail_view/'+str(payslip[0].pk))

def SalaryDeductions(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    emp_rep = models.Report.objects.filter(payroll_id=data.payroll_id)
    pay_id = data.payroll_id
    empData = []

    for emp in emp_rep:
        ded = models.Deductions.objects.filter(report=emp)
        if ded:
            for dedPay in ded:
                data = [emp.employee_id,emp.full_name,dedPay]
                empData.append(data)
    if 'pdf' in request.GET:
        return render_to_pdf_response(
            request,
            'payroll/salary_deductions_pdf.html',
            {'empData':empData},
            download_filename='Deductions.pdf',
            content_type='application/pdf',
        )
    else:
        return render(request, 'payroll/salary_deductions.html', {'empData':empData, 'pay_id':pay_id})

def SalaryAdditional(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    emp_rep = models.Report.objects.filter(payroll_id=data.payroll_id)
    pay_id = data.payroll_id
    empData = []
    total_amount = 0

    for emp in emp_rep:
        add = models.Additionalpays.objects.filter(report=emp)
        if add:
            for addPay in add:
                data = [emp.employee_id,emp.full_name,addPay]
                amount = addPay.amount
                total_amount += amount
                empData.append(data)
    if 'pdf' in request.GET:
        return render_to_pdf_response(
            request,
            'payroll/salary_additional_pdf.html',
            {'empData':empData},
            download_filename='AdditionalPay.pdf',
            content_type='application/pdf',
        )
    else:
        return render(request, 'payroll/salary_additional.html', {'empData':empData,'pay_id': pay_id, 'total_amount':total_amount,})



def EmpTimesheet(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    emp_rep = models.Report.objects.filter(payroll_id=data.payroll_id)
    pay_id = data.payroll_id
    timesheet = []

    for emp in emp_rep:
        time = models.ReportTimesheet.objects.filter(report=emp)
        timesheet.append([emp,time])
    return render(request, 'payroll/emp_timesheet.html', {'timesheet':timesheet, 'pay_id':pay_id,'emp':emp})


def load_workbook(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    emp = data.employee.all()
    #workbook initialized
    wb = Workbook()
    ws = wb.active

    SN=1
    #getting all cost center and work permit

    ws.append(['S.N','Employee ID','Name','Job Title','Department','Cost Center','Work Permit','Joining Date','Basic','Housing'
        ,'Transportation','Other','Gross','Other Income','Annual Ticket',
        'Salary Advance','Deduction','WLOP','NET Salary','WPS Cln1','WPS Cln2'])
    total_basic = 0
    total_housing = 0
    total_transport = 0
    total_other = 0
    total_gross = 0
    total_ot_in = 0
    total_an_ti = 0
    total_sa_in = 0
    total_ded = 0
    total_lwop = 0
    total_net_salary = 0
    total_wps1 = 0
    total_wps2 = 0
    for pays in emp:
        getPayslip = models.Payslip.objects.get(pk=pays.pk)
        empData = models.Employee.objects.get(pk=getPayslip.employee.pk)
        reports =  models.Report.objects.get(payroll_id=data.payroll_id,employee_id=empData.employee_id)
        getdedSum = models.Deductions.objects.filter(report=reports).aggregate(
            Sum('amount'))['amount__sum']
        dedSum = getdedSum or 0
        try:
            get_over_time = models.Additionalpays.objects.get(report=reports,additional='Overtime/ Holiday Pay')
            over_time = get_over_time.amount or 0
        except models.Additionalpays.DoesNotExist:
            over_time = 0
        try:
            getBonus = models.Additionalpays.objects.get(report=reports, additional='Bonus')
            bonus = getBonus.amount or 0
        except models.Additionalpays.DoesNotExist:
            bonus = 0
        try:
            get_reimbursement = models.Additionalpays.objects.get(report=reports,additional='Expense Reimbursement')
            reimbursement = get_reimbursement.amount or 0
        except models.Additionalpays.DoesNotExist:
            reimbursement = 0
        try:
            get_other = models.Additionalpays.objects.get(report=reports,additional='Other')
            other = get_other.amount or 0
        except models.Additionalpays.DoesNotExist:
            other = 0

        try:
            getAnnual_ticket = models.Additionalpays.objects.get(report=reports, additional='Annual Air Ticket')
            annual_ticket = getAnnual_ticket.amount or 0
        except models.Additionalpays.DoesNotExist:
            annual_ticket = 0
        try:
            getsalaryAdvance = models.Additionalpays.objects.get(report=reports, additional='Salary Advance')
            salary_income = getsalaryAdvance.amount or 0
        except models.Additionalpays.DoesNotExist:
            salary_income = 0
        other_income = over_time + bonus + reimbursement + other
        wps2 = other_income + annual_ticket + salary_income
        grossTotal = getPayslip.gross_pay + wps2 - dedSum
        net_salary = (getPayslip.gross_pay + other_income + annual_ticket + salary_income) - dedSum
        wps1 = reports.gross_pay - dedSum

        total_basic += reports.basic_salary
        total_housing += reports.housing_allowance
        total_transport += reports.transportation_allowance
        total_other += reports.other
        total_gross += reports.gross_pay
        total_ot_in += other_income
        total_an_ti += annual_ticket
        total_sa_in += salary_income
        total_ded += dedSum
        total_lwop += reports.lowp_deduction
        total_net_salary +=  net_salary
        total_wps1 += wps1
        total_wps2 += wps2

        ws.append([
            SN,
            empData.employee_id,
            empData.full_name,
            empData.designation,
            empData.department.department,
            reports.work_permit,
            reports.cost_center.cost_center,
            empData.joining_date,
            reports.basic_salary,
            reports.housing_allowance,
            reports.transportation_allowance,
            reports.other,
            reports.gross_pay,
            other_income,
            annual_ticket,
            salary_income,
            dedSum,
            reports.lowp_deduction,
            net_salary,
            wps1,
            wps2,
        ])
        SN += 1
    ws.append(['','','','','','','','Total',
        total_basic,
        total_housing,
        total_transport,
        total_other,
        total_gross,
        total_ot_in,
        total_an_ti,
        total_sa_in,
        total_ded,
        total_lwop,
        total_net_salary,
        total_wps1,
        total_wps2
    ])
    filename = "{}.xlsx".format(data.payroll_id)
    wb.save(filename)
    with open(filename, 'rb') as f:
        response = HttpResponse(
            f.read(),content_type='application/vnd.ms-excel'
        )
        response['Content-Disposition'] = 'inline;filename='+os.path.basename(filename)
    return response
    # return redirect('/payroll_detail/'+str(data.pk))

def payslipPdf(request,pk,empid):
    getPdf(pk,empid)
    return redirect('payroll:employee_list')

class AdditionalPDFView(PDFTemplateView):
    template_name = 'payroll/salary_additional.html'

class DeductionsPDFView(PDFTemplateView):
    template_name = 'payroll/salary_deductions.html'


def PaysDetailsPDFView(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = models.Report.objects.filter(payroll_id=data.payroll_id)
    datas = []
    total_empAdd = 0
    total_empDed = 0
    total_lwop = 0
    total_gross = 0
    total_net = 0

    for emp in  getEmp:
        empAdd = models.Additionalpays.objects.filter(report__employee_id=emp.employee_id, report__payroll_id=emp.payroll_id).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deductions.objects.filter(report__employee_id=emp.employee_id, report__payroll_id=emp.payroll_id).aggregate(Sum('amount'))['amount__sum']
        empTim = models.ReportTimesheet.objects.filter(report__employee_id=emp.employee_id, report__payroll_id=emp.payroll_id)
        empAdd = empAdd or 0
        empAdd = (round(empAdd,2))
        empDed = empDed or 0
        empDed = (round(empDed,2))
        gross = emp.gross_pay - emp.lowp_deduction

        net = gross + empAdd - empDed
        total_empAdd += empAdd
        total_empDed += empDed
        total_lwop += emp.lowp_deduction
        total_gross += gross
        total_net += net
        datas.append([emp,empAdd,empDed,empTim,gross,net])
    return render(request, 'payroll/pays_detail_pdf.html', {'datas':datas,'total_empAdd':total_empAdd,
                    'total_empDed':total_empDed,
                    'total_gross': total_gross,
                    'total_net':total_net,
                    'total_lwop':total_lwop,})

def ProcessSalaryAdditional(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = data.employee.all()
    addition = []
    total = 0
    for emp in getEmp:
        add = models.Additionalpay.objects.filter(payslip=emp)

        if add:
            for add1 in add:
                total += add1.amount
                addition.append([add1,emp])
    return render(request, 'payroll/process_salary_additional.html', {'addition':addition, 'total':total, 'data':data})

def ProcessSalaryDeductions(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = data.employee.all()
    deduction = []
    total = 0
    for emp in getEmp:
        ded = models.Deduction.objects.filter(payslip=emp)

        if ded:
            for ded1 in ded:
                total += ded1.amount
                deduction.append([ded1,emp])
    return render(request, 'payroll/process_salary_deductions.html', {'deduction':deduction, 'total':total, 'data':data})

def PayDetailListPDF(request):
    data = models.Payslip.objects.all()
    return render(request, 'payroll/pay_detail_list_pdf.html', {'data':data})


def ProcessSalaryTimesheet(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = data.employee.all()
    timesheet = []

    for emp in getEmp:
        time = models.Timesheet.objects.filter(employee=emp.employee)
        timesheet.append([emp,time])
    return render(request, 'payroll/process_salary_timesheet.html', {'timesheet':timesheet})


def ProcessSalaryPayDetails(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = data.employee.all()
    datas = []
    for emp in  getEmp:
        datas.append([emp])
    return render(request, 'payroll/process_salary_paydetails.html', {'datas':datas})


def ProcessSalaryMaster(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    master = []
    #getting all cost center and work permit
    cost_center = models.CostCenter.objects.all()
    work_permit = models.Payslip.objects.all().values_list('work_permit',flat=True).distinct()
    for cc in cost_center:
        for wp in work_permit:
            emp = models.ProcessSalary.objects.filter(
                    employee__work_permit=wp,
                    employee__cost_center=cc,
                    pk=data.pk
                    ).values_list("employee",flat=True)
            cc_wp = []
            if len(emp) > 0:
                total_basic = 0
                total_housing = 0
                total_transport = 0
                total_other = 0
                total_gross = 0
                total_ot_in = 0
                total_an_ti = 0
                total_sa_in = 0
                total_ded = 0
                total_lwops = 0
                total_net_salary = 0
                total_wps1 = 0
                total_wps2 = 0
                for pays in emp:
                    getPayslip = models.Payslip.objects.get(pk=pays)
                    empData = models.Employee.objects.get(pk=getPayslip.employee.pk)

                    getTimesheet = models.Timesheet.objects.get(employee=empData)
                    lwops = (getPayslip.gross_pay/getTimesheet.total)*getTimesheet.lwop
                    lwops = (round(lwops,2))

                    getdedSum = models.Deduction.objects.filter(payslip=getPayslip).aggregate(
                        Sum('amount'))['amount__sum']
                    dedSum = getdedSum or 0
                    dedSum = round(dedSum,2)
                    try:
                        get_over_time = models.Additionalpay.objects.get(payslip=getPayslip,additional='Overtime/ Holiday Pay')
                        over_time = get_over_time.amount or 0
                    except models.Additionalpay.DoesNotExist:
                        over_time = 0
                    try:
                        getBonus = models.Additionalpay.objects.get(payslip=getPayslip, additional='Bonus')
                        bonus = getBonus.amount or 0
                    except models.Additionalpay.DoesNotExist:
                        bonus = 0
                    try:
                        get_reimbursement = models.Additionalpay.objects.get(payslip=getPayslip,additional='Expense Reimbursement')
                        reimbursement = get_reimbursement.amount or 0
                    except models.Additionalpay.DoesNotExist:
                        reimbursement = 0
                    try:
                        get_other = models.Additionalpay.objects.get(payslip=getPayslip,additional='Other')
                        other = get_other.amount or 0
                    except models.Additionalpay.DoesNotExist:
                        other = 0

                    try:
                        getAnnual_ticket = models.Additionalpay.objects.get(payslip=getPayslip, additional='Annual Air Ticket')
                        annual_ticket = getAnnual_ticket.amount or 0
                    except models.Additionalpay.DoesNotExist:
                        annual_ticket = 0
                    try:
                        getsalaryAdvance = models.Additionalpay.objects.get(payslip=getPayslip, additional='Salary Advance')
                        salary_income = getsalaryAdvance.amount or 0
                    except models.Additionalpay.DoesNotExist:
                        salary_income = 0
                    other_income = over_time + bonus + reimbursement + other
                    wps2 = other_income + annual_ticket + salary_income
                    grossTotal = getPayslip.gross_pay + wps2 - dedSum
                    net_salary = (getPayslip.gross_pay + other_income + annual_ticket + salary_income) - dedSum - lwops
                    wps1 = getPayslip.gross_pay - dedSum

                    total_basic += getPayslip.basic_salary
                    total_housing += getPayslip.housing_allowance
                    total_transport += getPayslip.transportation_allowance
                    total_other += getPayslip.other
                    total_gross += getPayslip.gross_pay
                    total_ot_in += other_income
                    total_an_ti += annual_ticket
                    total_sa_in += salary_income
                    total_ded += dedSum
                    total_lwops += lwops
                    total_net_salary +=  net_salary
                    total_wps1 += wps1
                    total_wps2 += wps2

                    cc_wp.append([empData,getPayslip,other_income,annual_ticket,salary_income,dedSum,net_salary,wps1,wps2,lwops])
                master.append([cc.cost_center,wp,cc_wp,total_basic,total_housing,total_transport,
                    total_other,total_gross,total_ot_in,total_an_ti,total_sa_in,
                    total_ded,total_net_salary,total_wps1,total_wps2,data,total_lwops])

    return render(request, 'payroll/process_salary_master.html',{
        'master':master
        }
    )

def ProcessSalaryBank(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = data.employee.filter(payment_method="Bank")
    datas = []
    t_gross = 0
    t_vari = 0
    g_total = 0

    for emp in  getEmp:
        empAdd = models.Additionalpay.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deduction.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empTst = models.Timesheet.objects.get(employee=emp.employee)
        lwops = (emp.gross_pay/empTst.total)*empTst.lwop
        lwops = (round(lwops,2))
        empAddAmount = empAdd or 0
        empDedAmount = empDed or 0
        addamt = (round(empAddAmount,2))
        dedamt = (round(empDedAmount,2))
        vari = addamt-dedamt-lwops
        total = emp.gross_pay + vari
        t_gross +=emp.gross_pay
        t_vari += vari
        g_total += total
        datas.append([emp,vari,total,data,empTst])
    return render(request, 'payroll/process_salary_bank.html', {'datas':datas,'t_gross':t_gross,'t_vari':t_vari,'g_total':g_total,'data':data})

def ProcessSalaryLulu(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = data.employee.filter(payment_method="Lulu Exchange")
    datas = []
    t_gross = 0
    t_vari = 0
    g_total = 0

    for emp in  getEmp:
        empAdd = models.Additionalpay.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deduction.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empTst = models.Timesheet.objects.get(employee=emp.employee)
        lwops = (emp.gross_pay/empTst.total)*empTst.lwop
        lwops = (round(lwops,2))
        empAddAmount = empAdd or 0
        empDedAmount = empDed or 0
        addamt = (round(empAddAmount,2))
        dedamt = (round(empDedAmount,2))
        vari = addamt-dedamt-lwops
        total = emp.gross_pay + vari
        t_gross +=emp.gross_pay
        t_vari += vari
        g_total += total
        datas.append([emp,vari,total,empTst])

    return render(request, 'payroll/process_salary_lulu.html', {'datas':datas,'t_gross':t_gross,'t_vari':t_vari,'g_total':g_total,'data':data})

def load_workbook_bank(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = data.employee.filter(payment_method="Bank")
    wb = Workbook()
    ws = wb.active
    datas = []
    t_gross = 0
    t_vari = 0
    g_total = 0
    SN = 1
    ws.append(['S.N','Emplyee Unique ID','Agent ID','Employee Account With Agent','Income Fixed Component',
                    'Income Variable Component','Days on Leave without pay for the period',
                    'Pay Start Date','Pay End Date','Salary Payable Days For the period'])

    for emp in  getEmp:
        empAdd = models.Additionalpay.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deduction.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empTst = models.Timesheet.objects.get(employee=emp.employee)
        lwops = (emp.gross_pay/empTst.total)*empTst.lwop
        lwops = (round(lwops,2))
        empAddAmount = empAdd or 0
        empDedAmount = empDed or 0
        addamt = (round(empAddAmount,2))
        dedamt = (round(empDedAmount,2))
        vari = addamt-dedamt-lwops
        total = emp.gross_pay + vari
        t_gross +=emp.gross_pay
        t_vari += vari
        g_total += total
        ws.append([
            SN,
            emp.employee_unique,
            emp.agent_id,
            emp.bank_ac,
            emp.gross_pay,
            vari,
            empTst.lwop,
            data.start_date,
            data.finish_date,
            empTst.pay_days,
        ])
        SN += 1
    ws.append(['','','','Total',
                    t_gross,
                    t_vari,])
    filename = "{}.xlsx".format(data.payroll_id +"_"+"Bank_file")
    wb.save(filename)
    with open(filename, 'rb') as f:
        response = HttpResponse(
            f.read(),content_type='application/vnd.ms-excel'
        )
        response['Content-Disposition'] = 'inline;filename='+os.path.basename(filename)
    return response

def load_workbook_lulu(request,pk):
    data = models.ProcessSalary.objects.get(pk=pk)
    getEmp = data.employee.filter(payment_method="Lulu Exchange")
    wb = Workbook()
    ws = wb.active
    datas = []
    t_gross = 0
    t_vari = 0
    g_total = 0
    SN = 1

    ws.append(['COMPANY NAME: UNITEAM MEDIACAL ASSISTANCE LLC'])
    ws.append(['MOL ID: 600180'])
    ws.append(['SALARY MONTH:'+ str(data.salary_month)])

    ws.append(['S.N','NAME OF THE EMPLOYEE	','WORK PERMIT NO (8 DIGIT NO)','PERSONAL NO (14 DIGIT NO)','BANK NAME & ROUTING NO',
                    'BANK ACCOUNT NO','NO OF DAYS ABSENT',
                    'FIXED PORTION','VARIABLE PORTION	','TOTAL PAYMENT'])

    for emp in  getEmp:
        empAdd = models.Additionalpay.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deduction.objects.filter(payslip=emp).aggregate(Sum('amount'))['amount__sum']
        empTst = models.Timesheet.objects.get(employee=emp.employee)
        lwops = (emp.gross_pay/empTst.total)*empTst.lwop
        lwops = (round(lwops,2))
        empAddAmount = empAdd or 0
        empDedAmount = empDed or 0
        addamt = (round(empAddAmount,2))
        dedamt = (round(empDedAmount,2))
        vari = addamt-dedamt - lwops
        total = emp.gross_pay + vari
        t_gross +=emp.gross_pay
        t_vari += vari
        g_total += total
        ws.append([
            SN,
            emp.employee.full_name,
            emp.work_permit_8_digits,
            emp.personal_no_14_digits,
            emp.bank_name_routing,
            emp.bank_ac,
            empTst.lwop,
            emp.gross_pay,
            vari,
            total,

        ])
        SN += 1
    ws.append(['','','','','','','Total',
                    t_gross,
                    t_vari,
                    g_total,])
    filename = "{}.xlsx".format(data.payroll_id+"_"+"Lulu_file")
    wb.save(filename)
    with open(filename, 'rb') as f:
        response = HttpResponse(
            f.read(),content_type='application/vnd.ms-excel'
        )
        response['Content-Disposition'] = 'inline;filename='+os.path.basename(filename)
    return response

def SalaryRecord(request):
    employee = models.Employee.objects.all()
    return render(request, 'payroll/salary_record.html', {'employee_list':employee})

def SalaryRecordDetail(request,empid ):
    fullname = models.Employee.objects.get(employee_id=empid)
    record = models.Report.objects.filter(employee_id=empid)
    add = models.Additionalpays.objects.filter(report__employee_id=empid)
    deduction = models.Deductions.objects.filter(report__employee_id=empid)
    timesheet = models.ReportTimesheet.objects.filter(report__employee_id=empid)
    data = []
    total_lwop = 0
    total_gross = 0
    total_add = 0
    total_ded = 0
    total_net = 0

    for r in record:
        empAdd = models.Additionalpays.objects.filter(report=r).aggregate(Sum('amount'))['amount__sum']
        empDed = models.Deductions.objects.filter(report=r).aggregate(Sum('amount'))['amount__sum']
        empAdd = empAdd or 0
        empDed = empDed or 0
        empAdd = (round(empAdd,2))
        empDed = (round(empDed,2))
        gross = r.gross_pay - r.lowp_deduction
        net= gross + empAdd - empDed
        p_id = r.payroll_id

        total_lwop += r.lowp_deduction
        total_gross += gross
        total_add += empAdd
        total_ded += empDed
        total_net += net

        data.append([r,empAdd,empDed,net,gross,p_id])



    return render(request, 'payroll/salary_record_detail.html',
        {'record':record,
        'empid':empid,
        'add':add,
        'deduction':deduction,
        'timesheet':timesheet,
        'fullname':fullname,
        'data':data,
        'total_lwop':total_lwop,
        'total_gross': total_gross,
        "total_add": total_add,
        'total_ded': total_ded,
        'total_net': total_net,
        })

def downloadAll(request,pk):
    data = models.ProcessSalary.objects.get(payroll_id=pk)

    for emp in data.employee.all():
        filename = f"{emp.employee_id}__{pk}.pdf"
        filepath = settings.MEDIA_ROOT+"/pdf/"+filename
        if not os.path.exists(filepath):
            getPdf(pk,str(emp.employee_id))
            filepath = settings.MEDIA_ROOT+"/pdf/"+filename
    return redirect('payroll:payroll_register')

def email(request,pk):
    data = models.ProcessSalary.objects.get(payroll_id=pk)
    for emp in data.employee.all():
        filename = f"{emp.employee_id}__{pk}.pdf"
        filepath = settings.MEDIA_ROOT+"/pdf/"+filename
        if not os.path.exists(filepath):
            getPdf(pk,str(emp.employee_id))
            filepath = settings.MEDIA_ROOT+"/pdf/"+filename
        attachPath = filepath

        subject = 'Test email'
        message = 'Please ignore this email this is only for test purpose '
        email_from = settings.EMAIL_HOST_USER
        recipient_list = ['deepaksapkota1991@gmail.com']
        # recipient_list = [emp.employee.email]
        mail = EmailMessage( subject, message, email_from, recipient_list )
        mail.attach_file(attachPath)
        mail.send()
    return redirect('payroll:payroll_register')


def singleEmail(request,pk,empid):
    filename = empid+"__"+pk+".pdf"
    filepath = settings.MEDIA_ROOT+"/pdf/"+filename
    if not os.path.exists(filepath):
        getPdf(pk,empid)
        filepath = settings.MEDIA_ROOT+"/pdf/"+filename
    attachPath = filepath
    subject = 'Test email'
    message = 'Please ignore this email this is only for test purpose '
    email_from = settings.EMAIL_HOST_USER
    recipient_list = ['deepaksapkota1991@gmail.com']
    mail = EmailMessage( subject, message, email_from, recipient_list )
    mail.attach_file(attachPath)
    mail.send()
    return redirect('payroll:payroll_register')

class DisplayPdf(BaseDetailView):
    def get(self, request,*args, **kwargs):
        pk = self.kwargs.get('pk', None)
        empid = self.kwargs.get('empid', None)
        print(pk)
        print(empid)

        # for emp in data.employee.all():
        filename = f"{empid}__{pk}.pdf"
        filepath = settings.MEDIA_ROOT+"/pdf/"+filename
        if not os.path.exists(filepath):
            raise Http404
        response = FileResponse(open(filepath, 'rb'), content_type="application/pdf")
        response["Content-Disposition"] = "filename={}".format(filename)
        return response

def table(request):
    getPdf()
    return redirect('payroll:payroll_register')

def delete_everything(self):
    models.Timesheet.objects.all().delete()
    return redirect('payroll:timesheet_list')

def hr_records(request):
    return render(request,'payroll/hr_records.html')

class LabourContractListView(ListView):
    template_name = 'payroll/labour_contract_list.html'
    context_object_name = 'labour_contract_list'
    
    def get_queryset(self):
        queryset= models.LaborContract.objects.all()
        if "date" in self.request.GET:
            if self.request.GET["date"] !="":            
                queryset= models.LaborContract.objects.filter(contract_issue_date=self.request.GET["date"])
        if "hotel" in self.request.GET:
            if self.request.GET["hotel"] !="":            
                queryset= queryset.filter(employee__hotel=self.request.GET["hotel"])           
        return queryset

    def get_context_data(self, *args, **kwargs):
        context = super(LabourContractListView,self).get_context_data(*args,**kwargs)
        date = models.LaborContract.objects.all().values_list('contract_issue_date',flat=True).distinct()
        context['data'] = date
        hotel = models.Hotel.objects.all()       
        context['hotel'] = hotel
        search=''
        if self.request.GET:
            search="?date="+ self.request.GET['date'] +"&hotel="+self.request.GET['hotel']
        context['search'] = search
        return context

def LabourContractPrint(request):
    printlist = models.LaborContract.objects.order_by('employee__hotel__hotel','employee__first_name')
    c_date=""

    if "date" in request.GET:
        if request.GET["date"] !="":
            printlist = printlist.filter(contract_issue_date=request.GET["date"])
            c_date = request.GET["date"]
    if "hotel" in request.GET:
            if request.GET["hotel"] !="":            
                printlist= printlist.filter(employee__hotel=request.GET["hotel"])

    return render(request, 'payroll/labour_contract_print.html', {'print_list':printlist, 'c_date':c_date})

def LabourContractReport(request):
    printlist = models.LaborContract.objects.order_by('-contract_issue_date','employee__hotel__hotel','employee__first_name')

    return render(request, 'payroll/labour_contract_report.html', {'print_list':printlist})

class LabourContractDetailView(DetailView):
    model = models.LaborContract
    context_object_name = 'labour_contract_detail'
    template_name='payroll/labour_contract_detail.html'

    def get_context_data(self, *args, **kwargs):
        data = super(LabourContractDetailView,self).get_context_data(*args,**kwargs)
        return data
    
class LabourContractCreateView(CreateView):
    template_name = 'payroll/labour_contract_create.html'
    form_class = LabourContractForm
    success_url = reverse_lazy("payroll:labour_contract_list")

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

class LabourContractUpdateView(UpdateView):
    model = models.LaborContract
    form_class = LabourContractForm
    success_url = reverse_lazy("payroll:labour_contract_list")
    template_name='payroll/labour_contract_update.html'

class LabourContractDeleteView(DeleteView):
    model = models.LaborContract
    success_url = reverse_lazy("payroll:labour_contract_list")

class WorkPermitListView(ListView):
    template_name = 'payroll/work_permit_list.html'
    context_object_name = 'work_permit_list'

    def get_queryset(self):
        queryset= models.WorkPermit.objects.order_by('-contract_issue_date')
        if "date" in self.request.GET:
            if self.request.GET["date"] !="":            
                queryset= models.WorkPermit.objects.filter(contract_issue_date=self.request.GET["date"])
        if "hotel" in self.request.GET:
            if self.request.GET["hotel"] !="":            
                queryset= queryset.filter(employee__hotel=self.request.GET["hotel"])           
        return queryset

    def get_context_data(self, *args, **kwargs):
        context = super(WorkPermitListView,self).get_context_data(*args,**kwargs)
        date = models.WorkPermit.objects.all().order_by('contract_issue_date').values_list('contract_issue_date',flat=True).distinct()
        context['data'] = date
        hotel = models.Hotel.objects.all()       
        context['hotel'] = hotel
        search=''
        if self.request.GET:
            search="?date="+ self.request.GET['date'] +"&hotel="+self.request.GET['hotel']
        context['search'] = search
        return context

def WorkPermitPrint(request):
    printlist = models.WorkPermit.objects.order_by('employee__hotel__hotel')
    c_date=""
 

    if "date" in request.GET:
        if request.GET["date"] !="":
            printlist = printlist.filter(contract_issue_date=request.GET["date"])
            c_date = request.GET["date"]
    if "hotel" in request.GET:
            if request.GET["hotel"] !="":            
                printlist= printlist.filter(employee__hotel=request.GET["hotel"])

    return render(request, 'payroll/work_permit_print.html', {'print_list':printlist, 'c_date':c_date})

class WorkPermitDetailView(DetailView):
    model = models.WorkPermit
    context_object_name = 'work_permit_detail'
    template_name='payroll/work_permit_detail.html'

    def get_context_data(self, *args, **kwargs):
        data = super(WorkPermitDetailView,self).get_context_data(*args,**kwargs)
        return data

class WorkPermitCreateView(CreateView):
    template_name = 'payroll/work_permit_create.html'
    form_class = WorkPermitForm
    success_url = reverse_lazy("payroll:work_permit_list")

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

class WorkPermitUpdateView(UpdateView):
    model = models.WorkPermit
    form_class = WorkPermitForm
    success_url = reverse_lazy("payroll:work_permit_list")
    template_name='payroll/work_permit_update.html'

class WorkPermitDeleteView(DeleteView):
    model = models.WorkPermit
    success_url = reverse_lazy("payroll:work_permit_list")


class EidCollectListView(ListView):
    template_name = 'payroll/eid_collect_list.html'
    context_object_name = 'eid_collect_list'

    def get_queryset(self):
        queryset= models.EidCollect.objects.all()
        if "date" in self.request.GET:
            if self.request.GET["date"] !="":            
                queryset= models.EidCollect.objects.filter(eid_collected_date=self.request.GET["date"])
        if "hotel" in self.request.GET:
            if self.request.GET["hotel"] !="":            
                queryset= queryset.filter(employee__hotel=self.request.GET["hotel"])           
        return queryset

    def get_context_data(self, *args, **kwargs):
        context = super(EidCollectListView,self).get_context_data(*args,**kwargs)

        date = models.EidCollect.objects.all().values_list('eid_collected_date',flat=True).distinct()

        context['data'] = date
        hotel = models.Hotel.objects.all()       
        context['hotel'] = hotel
        search=''
        if self.request.GET:
            search="?date="+ self.request.GET['date'] +"&hotel="+self.request.GET['hotel']
        context['search'] = search
        return context

def EidCollectPrint(request):
    printlist = models.EidCollect.objects.order_by('employee__hotel__hotel')
    c_date=""

    if "date" in request.GET:
        if request.GET["date"] !="":
            printlist = printlist.filter(eid_collected_date=request.GET["date"])
            c_date = request.GET["date"]
    if "hotel" in request.GET:
            if request.GET["hotel"] !="":            
                printlist= printlist.filter(employee__hotel=request.GET["hotel"])

    return render(request, 'payroll/eid_collect_print.html', {'print_list':printlist, 'c_date':c_date})

class EidCollectDetailView(DetailView):
    model = models.EidCollect
    context_object_name = 'eid_collect_detail'
    template_name='payroll/eid_collect_detail.html'

    def get_context_data(self, *args, **kwargs):
        data = super(EidCollectDetailView,self).get_context_data(*args,**kwargs)
        return data

class EidCollectCreateView(CreateView):
    template_name = 'payroll/eid_collect_create.html'
    form_class = EidCollectForm
    success_url = reverse_lazy("payroll:eid_collect_list")

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

class EidCollectUpdateView(UpdateView):
    model = models.EidCollect
    form_class = EidCollectForm
    success_url = reverse_lazy("payroll:eid_collect_list")
    template_name='payroll/eid_collect_update.html'

class EidCollectDeleteView(DeleteView):
    model = models.EidCollect
    success_url = reverse_lazy("payroll:eid_collect_list")


class NewEidListView(ListView):
    template_name = 'payroll/new_eid_list.html'
    context_object_name = 'new_eid_list'

    def get_queryset(self):
        queryset= models.NewEid.objects.all()
        if "date" in self.request.GET:
            if self.request.GET["date"] !="":            
                queryset= models.NewEid.objects.filter(eid_received_date=self.request.GET["date"])
        if "hotel" in self.request.GET:
            if self.request.GET["hotel"] !="":            
                queryset= queryset.filter(employee__hotel=self.request.GET["hotel"])           
        return queryset

    def get_context_data(self, *args, **kwargs):
        context = super(NewEidListView,self).get_context_data(*args,**kwargs)
        date = models.NewEid.objects.all().values_list('eid_received_date',flat=True).distinct()
        context['data'] = date
        hotel = models.Hotel.objects.all()       
        context['hotel'] = hotel
        search=''
        if self.request.GET:
            search="?date="+ self.request.GET['date'] +"&hotel="+self.request.GET['hotel']
        context['search'] = search
        return context

def NewEidPrint(request):
    printlist = models.NewEid.objects.order_by('employee__hotel__hotel','employee__first_name')
    c_date=""

    if "date" in request.GET:
        if request.GET["date"] !="":
            printlist = printlist.filter(eid_received_date=request.GET["date"])
            c_date = request.GET["date"]
    if "hotel" in request.GET:
            if request.GET["hotel"] !="":            
                printlist= printlist.filter(employee__hotel=request.GET["hotel"])

    return render(request, 'payroll/new_eid_print.html', {'print_list':printlist, 'c_date':c_date})

class NewEidDetailView(DetailView):
    model = models.NewEid
    context_object_name = 'new_eid_detail'
    template_name='payroll/new_eid_detail.html'

    def get_context_data(self, *args, **kwargs):
        data = super(NewEidDetailView,self).get_context_data(*args,**kwargs)
        return data


class NewEidCreateView(CreateView):
    template_name = 'payroll/new_eid_create.html'
    form_class = NewEidForm
    success_url = reverse_lazy("payroll:new_eid_list")

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

class NewEidUpdateView(UpdateView):
    model = models.NewEid
    form_class = NewEidForm
    success_url = reverse_lazy("payroll:new_eid_list")
    template_name='payroll/new_eid_update.html'

class NewEidDeleteView(DeleteView):
    model = models.NewEid
    success_url = reverse_lazy("payroll:new_eid_list")

class DateWiseListView(ListView):
    template_name = 'payroll/labour_contract_list.html'
    context_object_name = 'labour_contract_list'
    model = models.LaborContract

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['date'] = models.LaborContract.objects.all().values_list('contract_issue_date',flat=True).distinct()
        return context

def neweid_upload(request):
    # declaring template
    template = "payroll/neweid_upload.html"
    # prompt is a context variable that can have different values      depending on their context
    prompt = {
        'order': 'Order of the CSV should be Employee id , basic_salary, housing_allowance, transportation_allowance, other, work_permit, cost_center, mode_of_payment, bank_ac, payment_method, employee_unique, agent_id, work_permit_8_digits, personal_no_14_digits, bank_name_routing'}
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        if column[0] != "":
            getemp = models.Employee.objects.get(first_name=column[0])           
            created = models.NewEid.objects.create(
            employee               =   getemp,
            eid_received_date      =   column[1],
            collected              =   column[2],
            )
    context = {}
    return redirect("new_eid_list")
# ### End of Bulk paydetail upload ###

def labourcontract_upload(request):
    # declaring template
    template = "payroll/labour_contract_upload.html"
    # prompt is a context variable that can have different values      depending on their context
    prompt = {
        'order': 'Order of the CSV should be Employee id , basic_salary, housing_allowance, transportation_allowance, other, work_permit, cost_center, mode_of_payment, bank_ac, payment_method, employee_unique, agent_id, work_permit_8_digits, personal_no_14_digits, bank_name_routing'}
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        if column[0] != "":
            getemp = models.Employee.objects.get(first_name=column[0])           
            created = models.LaborContract.objects.create(
            employee               =   getemp,
            # contract_issue_date    =   column[1],
            contract_collected     =   column[1],
            passport_collected     =   column[2],
            medical_done           =   column[3],
            passport_returned      =   column[4],
            )
    context = {}
    return redirect("labour_contract_list")
# ### End of Bulk paydetail upload ###

def passport_update(request):
    # declaring template
    template = "payroll/passport_update.html"

    # prompt is a context variable that can have different values depending on their context
    prompt = {
        'order': 'Order of the CSV should be Item , Category'
        }
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        try:
            employee = models.Employee.objects.get(first_name = column[0])
            try:
                passport = models.LaborContract.objects.get(employee = employee)
                passport.passport_returned = column[1]
                passport.passport_returned_date = column[2]
                passport.save()
            except models.LaborContract.DoesNotExist:
                print(column[0])
                pass
        except models.Employee.DoesNotExist:
            print(column[0])
            pass
    context = {}
    return render(request, template, context)

def employee_update(request):
    # declaring template
    template = "payroll/bulk_employee_update.html"

    # prompt is a context variable that can have different values depending on their context
    prompt = {
        'order': 'Order of the CSV should be Item , Category'
        }
    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template, prompt)
    csv_file = request.FILES['file']
    # let's check if it is a csv file
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'THIS IS NOT A CSV FILE')
    data_set = csv_file.read().decode('UTF-8')
    # setup a stream which is when we loop through each line we are able to handle a data in a stream
    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar="|"):
        try:
            employee = models.Employee.objects.get(employee_id = column[0])
            hotel    = models.Hotel.objects.get(hotel = column[1])
            employee.hotel = hotel
            employee.save()            
        except models.Employee.DoesNotExist:            
            pass
        except models.Employee.MultipleObjectsReturned:
            pass
    context = {}
    return render(request, template, context)


def p_contract_sign(request):
    printlist = models.LaborContract.objects.filter(contract_collected="False").exclude(employee__employee_status="Exiting")
    printlist  = printlist.order_by('-contract_issue_date','employee__hotel__hotel','employee__first_name')  
    return render(request, 'payroll/p_contract_sign.html', {'print_list':printlist})

def p_wp_sign(request):
    printlist = models.WorkPermit.objects.filter(contract_collected="False").exclude(employee__employee_status="Exiting")
    printlist  = printlist.order_by('-contract_issue_date', 'employee__hotel__hotel','employee__first_name')  
    return render(request, 'payroll/p_wp_sign.html', {'print_list':printlist})

def p_eid_collect(request):
    printlist = models.NewEid.objects.filter(collected="False").exclude(employee__employee_status="Exiting")
    printlist  = printlist.order_by('-eid_received_date', 'employee__hotel__hotel','employee__first_name')  
    return render(request, 'payroll/p_eid_collect.html', {'print_list':printlist})

def p_passport_collect(request):
    printlist = models.LaborContract.objects.filter(passport_collected="False").exclude(employee__employee_status="Exiting")
    printlist  = printlist.order_by('-contract_issue_date','employee__hotel__hotel','employee__first_name')
    return render(request, 'payroll/p_passport_collect.html', {'print_list':printlist})

def p_visa_medical(request):
    printlist = models.LaborContract.objects.filter(medical_done="False").exclude(employee__employee_status="Exiting")
    printlist  = printlist.order_by('-contract_issue_date','employee__hotel__hotel','employee__first_name')    
    return render(request, 'payroll/p_visa_medical.html', {'print_list':printlist})

def p_p_collect(request):
    printlist = models.LaborContract.objects.filter(~Q(employee__employee_status="Exiting") & ~Q(passport_receive_date=None) & Q(passport_returned=False)) 
    printlist  = printlist.order_by('-passport_receive_date','employee__hotel__hotel','employee__first_name')    
    return render(request, 'payroll/p_p_collect.html', {'print_list':printlist})


class PassportReturnListView(ListView):
    template_name = 'payroll/p_return_list.html'
    context_object_name = 'p_return_list'

    
    def get_queryset(self):
        queryset= models.LaborContract.objects.all()
        if "date" in self.request.GET:
            if self.request.GET["date"] !="":            
                queryset= models.LaborContract.objects.filter(passport_receive_date=self.request.GET["date"])
        if "hotel" in self.request.GET:
            if self.request.GET["hotel"] !="":            
                queryset= queryset.filter(employee__hotel=self.request.GET["hotel"])           
        return queryset

    def get_context_data(self, *args, **kwargs):
        context = super(PassportReturnListView,self).get_context_data(*args,**kwargs)
        date = models.LaborContract.objects.all().values_list('passport_receive_date',flat=True).distinct().order_by('passport_receive_date')
        context['data'] = date
        hotel = models.Hotel.objects.all()       
        context['hotel'] = hotel
        search=''
        if self.request.GET:
            search="?date="+ self.request.GET['date'] +"&hotel="+self.request.GET['hotel']
        context['search'] = search
        return context

def PassportReturnPrint(request):
    printlist = models.LaborContract.objects.order_by('employee__hotel__hotel','employee__first_name')
    c_date=""
    if "date" in request.GET:
        if request.GET["date"] !="":
            printlist = printlist.filter(passport_receive_date=request.GET["date"])
            c_date = request.GET["date"]
    if "hotel" in request.GET:
            if request.GET["hotel"] !="":            
                printlist= printlist.filter(employee__hotel=request.GET["hotel"])
    return render(request, 'payroll/p_return_print.html', {'print_list':printlist, 'c_date':c_date})

class PassportReturnUpdateView(UpdateView):
    model = models.LaborContract
    form_class = PassportReturnUpdateForm
    success_url = reverse_lazy("payroll:p_return_list")
    template_name='payroll/p_return_update.html'

def multiple_passport_return(request):
    labour_contracte_list = models.LaborContract.objects.all().values_list('employee',flat=True)
    employee_list          = models.Employee.objects.filter(pk__in=labour_contracte_list).exclude(employee_status="Exiting")
    if request.method == 'POST':
        employees               = request.POST.getlist('employees[]')
        passport_receive_date   = request.POST.get('passport_receive_date')
        passport_return         = request.POST.get('passport_return')
        if passport_return == 'on':
            passport_return = True
        else:
            passport_return = False
        return_to_employee      = request.POST.get('return_to_employee')
        note                    = request.POST.get('note')
        selected_employee       = models.Employee.objects.filter(employee_id__in=employees)        
        for emp in selected_employee:
            labour_contract = models.LaborContract.objects.get(employee__employee_id=emp.employee_id)
            if passport_receive_date !="":
                labour_contract.passport_receive_date = passport_receive_date                
            labour_contract.passport_returned = passport_return
            if return_to_employee !="":
                labour_contract.passport_returned_date = return_to_employee
            labour_contract.save()            
    return render(request, 'payroll/m_passport_return_update.html', {'employee_list':employee_list})

def multiple_passport_collect(request):
    labour_contracte_list = models.LaborContract.objects.all().values_list('employee',flat=True)
    employee_list          = models.Employee.objects.filter(pk__in=labour_contracte_list).exclude(employee_status="Exiting")
    if request.method == 'POST':
        employees               = request.POST.getlist('employees[]')
        passport_return         = request.POST.get('passport_return')
        if passport_return == 'on':
            passport_return = True
        else:
            passport_return = False
        return_to_employee      = request.POST.get('return_to_employee')
        note                    = request.POST.get('note')
        selected_employee       = models.Employee.objects.filter(employee_id__in=employees)        
        for emp in selected_employee:
            labour_contract = models.LaborContract.objects.get(employee__employee_id=emp.employee_id)
            labour_contract.passport_returned = passport_return
            if return_to_employee !="":
                labour_contract.passport_returned_date = return_to_employee
            labour_contract.save()            
    return render(request, 'payroll/m_passport_return_collected.html', {'employee_list':employee_list})

def labour_contract_create_multiple(request):
    employee_list         = models.Employee.objects.exclude(employee_status="Exiting")    
    if request.method == 'POST':
        employees               = request.POST.getlist('employees[]')
        contract_issued_date   = request.POST.get('contract_issued_date')
        selected_employee       = models.Employee.objects.filter(employee_id__in=employees)  
        messages.success(request, 'Your Data bas been submitted successfully')      
        for emp in selected_employee:
            labour_contract = models.LaborContract(employee=emp, contract_issue_date = contract_issued_date)
            labour_contract.save()
    return render(request, 'payroll/m_labour_contract_create.html',{'employee_list':employee_list})

def new_eid_create_multiple(request):
    employee_list         = models.Employee.objects.exclude(employee_status="Exiting")    
    if request.method == 'POST':
        employees               = request.POST.getlist('employees[]')
        eid_received_date   = request.POST.get('eid_received_date')
        selected_employee       = models.Employee.objects.filter(employee_id__in=employees)
        messages.success(request, 'Your Data bas been submitted successfully')
              
        for emp in selected_employee:
            new_eid = models.NewEid(employee=emp, eid_received_date = eid_received_date)
            new_eid.save()
    return render(request, 'payroll/m_new_eid_create.html',{'employee_list':employee_list})

def multiple_new_eid_return(request):
    new_eid_list           = models.NewEid.objects.all().values_list('employee',flat=True)
    employee_list          = models.Employee.objects.filter(pk__in=new_eid_list).exclude(employee_status="Exiting")
    if request.method == 'POST':
        employees               = request.POST.getlist('employees[]')
        collected               = request.POST.get('collected')
        if collected == 'on':
            collected = True
        else:
            collected = False
        eid_collected_date      = request.POST.get('eid_collected_date')
        note                    = request.POST.get('note')       
        selected_employee       = models.Employee.objects.filter(employee_id__in=employees)        
        for emp in selected_employee:
            new_eid = models.NewEid.objects.get(employee__employee_id=emp.employee_id)
            new_eid.collected = collected
            new_eid.eid_collected_date = eid_collected_date            
            new_eid.note = note
            new_eid.save()            
    return render(request, 'payroll/m_new_eid_update.html', {'employee_list':employee_list})

def employee_export(request):
    data = models.Employee.objects.all()
    wb = Workbook()
    ws = wb.active
    SN = 1
    ws.append(['S.N','Aspen ID','Name','Phone','Hotel', 'Employee Status','Contract Issue Date',
                'Contract Collected','Contract Collected Date','Passport Collected','Passport Collected Date',
                'Medical','Medical Date','Passport Receive Date','Passport Returned ?','Passport Returned Date',
                'EID Receive Date','EID collected Date','Work Permit Issued Date','Work Permit Collected Date',])
    for employee in data:
        eid_receive = None
        eid_collected_date = None
        try:
            emp = models.LaborContract.objects.get(employee__employee_id=employee.employee_id)            
            try:
                new_eid = models.NewEid.objects.get(employee__employee_id=employee.employee_id)
                eid_receive = new_eid.eid_received_date
                eid_collected_date = new_eid.eid_collected_date
            except models.NewEid.DoesNotExist:
                pass
            ws.append([
                SN,
                employee.employee_id,
                employee.first_name,
                employee.phone,
                employee.hotel.hotel, 
                employee.employee_status,
                emp.contract_issue_date,
                emp.contract_collected,
                emp.contract_collected_date,
                emp.passport_collected,
                emp.passport_collect_date,
                emp.medical_done,
                emp.medical_date,
                emp.passport_receive_date,
                emp.passport_returned,
                emp.passport_returned_date,
                eid_receive,
                eid_collected_date,
                '',''
            ])            
        except models.LaborContract.DoesNotExist:
            try:
                w_p = models.WorkPermit.objects.get(employee__employee_id=employee.employee_id)
                ws.append([
                    SN,
                    employee.employee_id,
                    employee.first_name,
                    employee.phone,
                    employee.hotel.hotel, 
                    employee.employee_status,
                    '','','','','','','','','','','','',
                    w_p.contract_issue_date,
                    w_p.contract_collected_date
                    
                ])
            except models.WorkPermit.DoesNotExist:
                ws.append([
                    SN,
                    employee.employee_id,
                    employee.first_name,
                    employee.phone,
                    employee.hotel.hotel, 
                    employee.employee_status,
                    '','','','','','','','','','','','','',''            
                ])
        except models.LaborContract.MultipleObjectsReturned:
            continue           
        SN += 1
    filename = "HR_Records.xlsx"
    wb.save(filename)
    with open(filename, 'rb') as f:
        response = HttpResponse(
            f.read(),content_type='application/vnd.ms-excel'
        )
        response['Content-Disposition'] = 'inline;filename='+os.path.basename(filename)
    return response